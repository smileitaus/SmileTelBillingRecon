import openpyxl
import json

wb = openpyxl.load_workbook('/home/ubuntu/billing-tool/SM.xlsx', data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    
    # Print first 5 rows
    rows = list(ws.iter_rows(values_only=True))
    
    # Find header row (first non-empty row)
    header_row = None
    for i, row in enumerate(rows[:10]):
        non_empty = [c for c in row if c is not None and str(c).strip() != '']
        if len(non_empty) >= 3:
            header_row = i
            break
    
    if header_row is not None:
        print(f"Header row index: {header_row}")
        headers = rows[header_row]
        print(f"Headers: {list(headers)}")
        
        # Print first 10 data rows
        print(f"\nFirst 10 data rows:")
        for row in rows[header_row+1:header_row+11]:
            if any(c is not None and str(c).strip() != '' for c in row):
                print(f"  {list(row)}")
        
        # Count total data rows
        data_rows = [r for r in rows[header_row+1:] if any(c is not None and str(c).strip() != '' for c in r)]
        print(f"\nTotal data rows: {len(data_rows)}")
        
        # Identify key columns
        print(f"\nColumn analysis:")
        for i, h in enumerate(headers):
            if h is not None:
                col_vals = [rows[r][i] for r in range(header_row+1, min(header_row+20, len(rows))) if rows[r][i] is not None]
                print(f"  Col {i} '{h}': sample={col_vals[:3]}")
    else:
        print("Could not find header row")
        print("First 5 rows:")
        for row in rows[:5]:
            print(f"  {list(row)}")
